"""L2 — Integration tests. Services combined, not mocked.

Three flows the architecture claims to support end to end: a conversation about
an uploaded document, a tool run that produces a workspace, and a client that
loses its connection mid-stream.
"""
from __future__ import annotations

import io
import json

import pytest
from conftest import run_turn, wait_for_turn, wait_until

from primnox2.assets import service as assets
from primnox2.chat import turns
from primnox2.kernel.events import bus
from primnox2.sandbox import manager as sandbox
from primnox2.storage import db
from primnox2.workspaces import service as workspaces

PDF_SENTENCE = "Primnox stores every uploaded file as a content addressed asset."


def _make_pdf(text: str = PDF_SENTENCE) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    c.setFont("Helvetica", 14)
    c.drawString(72, 700, text)
    c.showPage()
    c.save()
    return buf.getvalue()


# ── PDF conversation ─────────────────────────────────────────────────────────
class TestPdfConversation:
    def test_upload_ingest_context_chat(self, conversation, events, scripted):
        pdf = _make_pdf()
        turn = turns.create_turn(conversation, "summarise the attached pdf")
        tid = turn["turn_id"]

        asset = assets.ingest_bytes(pdf, "primnox.pdf",
                                    conversation_id=conversation, turn_id=tid)
        wait_until(lambda: assets.get(asset["id"])["status"] == "ready",
                   what="pdf ingestion")

        stored = assets.get(asset["id"])
        assert stored["kind"] == "pdf"
        assert stored["page_count"] == 1
        assert "content addressed" in (stored["extracted_text"] or ""), \
            "text was not extracted from the pdf"

        # Exactly one asset for these bytes.
        rows = db.connect().execute(
            "SELECT COUNT(*) AS n FROM assets WHERE sha256=?", (stored["sha256"],)).fetchone()
        assert rows["n"] == 1

        # The turn references it.
        assert [a["id"] for a in assets.for_turn(tid)] == [asset["id"]]

        # And the chat can actually answer from it.
        from primnox2.context import service as context
        bundle = context.build(conversation, "what does it say?", turn_id=tid, budget=50_000)
        joined = "\n".join(m["content"] for m in bundle.messages)
        assert "content addressed" in joined
        assert "primnox.pdf" in joined, "citation source was not preserved"

    def test_no_duplicate_ingestion(self, conversation):
        pdf = _make_pdf("A document uploaded twice is still one document.")
        first = assets.ingest_bytes(pdf, "twice.pdf", conversation_id=conversation)
        wait_until(lambda: assets.get(first["id"])["status"] == "ready", what="ingestion")

        jobs_before = db.connect().execute(
            "SELECT COUNT(*) AS n FROM jobs WHERE kind='asset.ingest'").fetchone()["n"]

        second = assets.ingest_bytes(pdf, "twice-copy.pdf", conversation_id=conversation)
        assert second["id"] == first["id"], "identical bytes produced a second asset"
        assert second["deduplicated"] is True

        jobs_after = db.connect().execute(
            "SELECT COUNT(*) AS n FROM jobs WHERE kind='asset.ingest'").fetchone()["n"]
        assert jobs_after == jobs_before, "a duplicate upload queued a second ingestion"

    def test_turn_waits_for_unready_asset(self, conversation, scripted):
        """CRS §2.6 — a turn must never proceed with empty content."""
        scripted("I have read it.")
        turn = turns.create_turn(conversation, "read this")
        tid = turn["turn_id"]

        # Register an asset and hold it in `ingesting`, as a slow OCR would.
        #
        # Waiting for ingestion to finish before forcing the status back is the
        # whole reason this is deterministic: `ingest_bytes` queues a job, and
        # setting `ingesting` while that job is still in flight is a race the
        # ingest worker wins about half the time — it writes `ready` after the
        # test wrote `ingesting`, the turn sees nothing pending, and the test
        # times out waiting for a park that correctly never happened.
        asset = assets.ingest_bytes(b"slow document bytes", "slow.txt",
                                    conversation_id=conversation, turn_id=tid)
        wait_until(lambda: assets.get(asset["id"])["status"] == "ready",
                   what="the ingest job to finish before it is held back")
        with db.tx() as c:
            c.execute("UPDATE assets SET status='ingesting' WHERE id=?", (asset["id"],))

        from primnox2.kernel import scheduler as sched
        sched.enqueue(tid, "chat.reply", {"conversation_id": conversation, "text": "read this"})

        wait_until(lambda: db.connect().execute(
            "SELECT status FROM turns WHERE id=?", (tid,)).fetchone()["status"] == "tool_running",
            what="turn to park while the asset ingests")

        with db.tx() as c:
            c.execute("UPDATE assets SET status='ready' WHERE id=?", (asset["id"],))
        assert wait_for_turn(tid) == "completed"


# ── Tool integration ─────────────────────────────────────────────────────────
class TestToolIntegration:
    def test_generate_execute_workspace(self, conversation, events, scripted, sandbox_ready):
        """Generate Python → execute → create a workspace, in one turn."""
        scripted(
            '<tool name="run_python">\n{"code": "print(6*7)"}\n</tool>',
            '<tool name="create_workspace">\n'
            '{"kind": "python", "title": "Answer", "files": {"main.py": "print(6*7)"}}\n'
            '</tool>',
            "Done — the answer is 42 and I saved the script.",
        )
        tid = run_turn(conversation, "compute 6*7 and save the script")
        assert wait_for_turn(tid, timeout=180) == "completed"

        # One chat job owns the whole turn.
        jobs = db.connect().execute(
            "SELECT kind FROM jobs WHERE turn_id=?", (tid,)).fetchall()
        assert [j["kind"] for j in jobs] == ["chat.reply"], \
            "the turn's work was split across jobs instead of owned by one"

        # One execution session, recorded and addressable.
        sessions = sandbox.for_turn(tid)
        assert len(sessions) == 1, f"expected one execution session, got {len(sessions)}"
        session = sandbox.get(sessions[0]["id"])
        assert session["status"] in ("completed", "destroyed")
        assert session["exit_code"] == 0
        assert "42" in (session["stdout"] or ""), "execution output was not attached"
        assert session["snapshot"] is not None, "no snapshot was taken"

        # The workspace exists and is versioned.
        created = workspaces.for_turn(tid)
        assert len(created) == 1
        ws = workspaces.get(created[0]["id"])
        assert ws["files"]["main.py"] == "print(6*7)"
        assert ws["current_version"] == 1

        # And the whole thing is visible on the event stream.
        kinds = events.kinds(tid)
        for expected in ("tool.call", "sandbox.created", "sandbox.completed",
                         "tool.result", "workspace.created", "turn.completed"):
            assert expected in kinds, f"{expected} never reached the client"

    def test_tool_failure_does_not_fail_the_turn(self, conversation, events, scripted,
                                                 sandbox_ready):
        """CRS §6.2.2 — a job's failure must not silently fail its turn."""
        scripted(
            '<tool name="run_python">\n{"code": "raise SystemExit(3)"}\n</tool>',
            "That script exited with an error, here is what happened.",
        )
        tid = run_turn(conversation, "run something broken")
        assert wait_for_turn(tid, timeout=180) == "completed"

        results = events.of_kind("tool.result", tid)
        assert results and results[0]["payload"]["status"] == "error"

    def test_unknown_tool_is_correctable_not_fatal(self, conversation, events, scripted):
        scripted(
            '<tool name="make_coffee">\n{"strength": "strong"}\n</tool>',
            "Sorry, I cannot do that. Here is a real answer instead.",
        )
        tid = run_turn(conversation, "make coffee")
        assert wait_for_turn(tid) == "completed"
        assert "real answer" in turns.get_history(conversation)[-1]["assistant_message"]["text"]

    def test_malformed_tool_block_gets_one_correction(self, conversation, scripted):
        scripted(
            '<tool name="run_python">\n{this is not json}\n</tool>',
            "Corrected — here is the answer.",
        )
        tid = run_turn(conversation, "do a thing")
        assert wait_for_turn(tid) == "completed"


# ── Streaming, disconnect and replay ─────────────────────────────────────────
class ClientFold:
    """A minimal CRS client (§8.4): dedupe by event_id, buffer out-of-order,
    pure fold. The frontend's reducer must behave identically."""

    def __init__(self, cursor: int = 0) -> None:
        self.cursor = cursor
        self.seen: set[str] = set()
        self.text: dict[str, str] = {}
        self.buffer: dict[int, dict] = {}

    def deliver(self, event: dict) -> None:
        if event.get("sequence") is None:
            return
        if event["event_id"] in self.seen:
            return                                   # §8.4.1 dedupe
        if event["sequence"] > self.cursor + 1:
            self.buffer[event["sequence"]] = event    # §8.4.2 buffer the gap
            return
        self._apply(event)
        while self.cursor + 1 in self.buffer:
            self._apply(self.buffer.pop(self.cursor + 1))

    def _apply(self, event: dict) -> None:
        self.seen.add(event["event_id"])
        self.cursor = max(self.cursor, event["sequence"])
        if event["kind"] == "token":
            self.text[event["turn_id"]] = self.text.get(event["turn_id"], "") + event["payload"]["text"]


class TestStreamingAndReplay:
    def test_disconnect_reconnect_replays_only_the_gap(self, conversation, scripted):
        scripted("One two three four five six seven eight.", chunk=3)

        live: list[dict] = []
        sid = bus.subscribe(live.append)
        try:
            tid = run_turn(conversation, "count")
            assert wait_for_turn(tid) == "completed"
        finally:
            bus.unsubscribe(sid)

        turn_events = [e for e in live if e.get("turn_id") == tid and e.get("sequence")]
        assert len(turn_events) > 3, "not enough events to simulate a mid-stream drop"

        # The client saw the first few, then dropped.
        split = len(turn_events) // 2
        client = ClientFold(cursor=turn_events[0]["sequence"] - 1)
        for e in turn_events[:split]:
            client.deliver(e)
        cursor_at_drop = client.cursor

        replayed = bus.replay(cursor_at_drop, [conversation])
        assert replayed, "reconnect replayed nothing despite a known gap"
        assert all(e["sequence"] > cursor_at_drop for e in replayed), \
            "replay re-sent events the client already had"

        for e in replayed:
            client.deliver(e)

        final = turns.get_history(conversation)
        stored = [t for t in final if t["turn_id"] == tid][0]["assistant_message"]["text"]
        assert client.text[tid] == stored, "the reconnected client's text differs from stored state"

    def test_duplicate_delivery_is_idempotent(self, conversation, scripted):
        scripted("Exactly once, please.", chunk=4)
        live: list[dict] = []
        sid = bus.subscribe(live.append)
        try:
            tid = run_turn(conversation, "say it once")
            wait_for_turn(tid)
        finally:
            bus.unsubscribe(sid)

        turn_events = [e for e in live if e.get("turn_id") == tid and e.get("sequence")]
        client = ClientFold(cursor=turn_events[0]["sequence"] - 1)
        for e in turn_events:
            client.deliver(e)
            client.deliver(e)          # every packet delivered twice
        assert client.text[tid] == "Exactly once, please."

    def test_out_of_order_delivery_is_buffered(self, conversation, scripted):
        scripted("Order matters here.", chunk=3)
        live: list[dict] = []
        sid = bus.subscribe(live.append)
        try:
            tid = run_turn(conversation, "order")
            wait_for_turn(tid)
        finally:
            bus.unsubscribe(sid)

        turn_events = [e for e in live if e.get("turn_id") == tid and e.get("sequence")]
        client = ClientFold(cursor=turn_events[0]["sequence"] - 1)
        shuffled = list(reversed(turn_events))       # worst case: fully reversed
        for e in shuffled:
            client.deliver(e)
        assert client.text[tid] == "Order matters here.", \
            "out-of-order packets were applied out of order"

    def test_caught_up_client_replays_nothing(self, conversation, scripted):
        """CRS §8.1.1."""
        scripted("Already current.")
        tid = run_turn(conversation, "hi")
        wait_for_turn(tid)
        assert bus.replay(bus.head(), [conversation]) == []

    def test_replay_filtered_to_open_conversations(self, conversation, scripted):
        """§8.2.1 — and safe only because history never comes from the log."""
        other = turns.create_conversation("Unrelated")["id"]
        scripted("In the other room.")
        before = bus.head()
        tid = run_turn(other, "hello")
        wait_for_turn(tid)

        replayed = bus.replay(before, [conversation])
        assert all(e["conversation_id"] != other for e in replayed)


# ── Incognito ────────────────────────────────────────────────────────────────
TABLES = ("conversations", "turns", "messages", "events", "jobs", "assets",
          "asset_chunks", "turn_assets", "workspaces", "workspace_versions",
          "execution_sessions")


def _row_counts() -> dict[str, int]:
    conn = db.connect()
    return {t: conn.execute(f"SELECT COUNT(*) AS n FROM {t}").fetchone()["n"]
            for t in TABLES}


class TestIncognito:
    """CRS §11.2 — a conversation that writes nothing.

    The assertion that matters is the row count, taken across every table a
    turn could plausibly touch. Checking `conversations` alone would pass while
    the user's message sat in `jobs.payload` and their reply sat in `events`,
    which is exactly how this could be got wrong.
    """

    @pytest.fixture(autouse=True)
    def _clean(self):
        from primnox2.chat import ephemeral

        yield
        ephemeral.reset()

    def test_a_whole_turn_writes_nothing(self, scripted, events):
        scripted("Nothing about this is written down.")
        before = _row_counts()

        conv = turns.create_conversation("Private", incognito=True)
        assert conv["incognito"] is True
        tid = run_turn(conv["id"], "remember the passphrase: correct horse battery")
        assert wait_for_turn(tid) == "completed"

        assert _row_counts() == before, "an incognito turn wrote to the database"

        # And not by being silently dropped — the turn ran, and the client saw
        # every stage of it.
        kinds = events.kinds(tid)
        for expected in ("turn.created", "turn.status", "token", "turn.completed"):
            assert expected in kinds, f"{expected} never reached the client"
        assert events.text(tid) == "Nothing about this is written down."

    def test_its_events_carry_no_sequence(self, scripted, events):
        """§11.2.2 — in memory, delivered live, never in the log. A sequence
        number would also punch a hole in every other client's replay."""
        scripted("Ephemeral.")
        head_before = bus.head()

        conv = turns.create_conversation("Private", incognito=True)
        tid = run_turn(conv["id"], "hello")
        wait_for_turn(tid)

        assert bus.head() == head_before, "an incognito turn advanced the global cursor"
        assert all(e["sequence"] is None for e in events.for_turn(tid))
        assert bus.replay(head_before, [conv["id"]]) == []

    def test_history_survives_a_reload_but_not_a_restart(self, scripted):
        """The transcript lives as long as the process and no longer."""
        from primnox2.chat import ephemeral

        scripted("First.", "Second.")
        conv = turns.create_conversation("Private", incognito=True)
        wait_for_turn(run_turn(conv["id"], "one"))
        wait_for_turn(run_turn(conv["id"], "two"))

        # A reload is a history read, and it finds both turns.
        history = turns.get_history(conv["id"])
        assert [t["seq"] for t in history] == [1, 2]
        assert history[0]["assistant_message"]["text"] == "First."

        # A restart is the store going away with the process.
        ephemeral.reset()
        assert turns.get_history(conv["id"]) == []
        assert not turns.is_incognito(conv["id"])

    def test_the_second_turn_can_see_the_first(self, scripted):
        """History has to reach the model, or an incognito chat is a series of
        strangers rather than a conversation."""
        from primnox2.context import service as context

        scripted("Noted.")
        conv = turns.create_conversation("Private", incognito=True)
        wait_for_turn(run_turn(conv["id"], "my code name is Bluebird"))

        bundle = context.build(conv["id"], "what is my code name?", budget=50_000)
        joined = "\n".join(m["content"] for m in bundle.messages)
        assert "Bluebird" in joined, "the earlier turn never reached the prompt"
        assert "Noted." in joined

    def test_tools_that_persist_are_refused_not_run(self, scripted, events):
        """§11.2.4 — refusing beats quietly writing a workspace to disk."""
        scripted(
            '<tool name="create_workspace">\n'
            '{"kind": "python", "title": "Leak", "files": {"main.py": "print(1)"}}\n'
            '</tool>',
            "I can't do that in an incognito chat.",
        )
        before = _row_counts()
        conv = turns.create_conversation("Private", incognito=True)
        tid = run_turn(conv["id"], "save this script")
        assert wait_for_turn(tid) == "completed"

        assert _row_counts() == before, "a refused tool still wrote something"
        assert workspaces.for_turn(tid) == []

    def test_an_ordinary_conversation_still_persists(self, scripted):
        """The guard rail: none of the above may leak into the normal path."""
        scripted("Written down.")
        conv = turns.create_conversation("Ordinary")
        assert conv["incognito"] is False
        tid = run_turn(conv["id"], "keep this")
        assert wait_for_turn(tid) == "completed"

        row = db.connect().execute("SELECT status FROM turns WHERE id=?", (tid,)).fetchone()
        assert row["status"] == "completed"
        text = db.connect().execute(
            "SELECT text FROM messages WHERE turn_id=? AND role='assistant'",
            (tid,)).fetchone()["text"]
        assert text == "Written down."

    def test_retrying_an_incognito_turn_starts_a_new_one(self, scripted):
        """Retry re-read the new turn from the table to find out what to run,
        which for an incognito turn found nothing: the retry raised and left a
        turn queued that no worker would ever pick up."""
        import asyncio

        from primnox2 import app as app_module

        scripted("Second time lucky.")
        before = _row_counts()
        conv = turns.create_conversation("Private", incognito=True)
        first = turns.create_turn(conv["id"], "try this")
        turns.fail(first["turn_id"], "provider_unreachable", "down", True)

        retried = asyncio.run(app_module.retry_turn(first["turn_id"]))
        assert retried["turn_id"] != first["turn_id"], "retry reopened the failed turn"
        assert wait_for_turn(retried["turn_id"]) == "completed"

        history = turns.get_history(conv["id"])
        assert [t["status"] for t in history] == ["failed", "completed"]
        assert _row_counts() == before, "the retried turn wrote to the database"


class TestMultiStepReplyText:
    def test_two_model_calls_do_not_run_together(self, conversation, scripted,
                                                 sandbox_ready):
        """Found live: a turn whose two calls both ended with "144" was stored
        as "14412 * 12 = 144" — no separator between one call's prose and the
        next, so the last word of one ran into the first of the other."""
        scripted(
            '12 * 12 = 144<tool name="run_python">\n{"code": "print(12*12)"}\n</tool>',
            "12 * 12 = 144",
        )
        tid = run_turn(conversation, "compute 12*12")
        assert wait_for_turn(tid, timeout=180) == "completed"

        text = db.connect().execute(
            "SELECT text FROM messages WHERE turn_id=? AND role='assistant'",
            (tid,)).fetchone()["text"]
        assert "14412" not in text, f"two calls ran together: {text!r}"
        assert text.count("144") == 2


class TestWeakModelFailureModes:
    """Both of these were produced by qwen2.5:7b, not imagined."""

    def test_an_unfinished_tool_block_fails_rather_than_replying_blank(
            self, conversation, scripted):
        """2 of 8 turns completed with an empty assistant message: the model
        opened a tool block and never closed it, the filter correctly withheld
        the markup, and there was no prose behind it. A turn that finishes
        having said nothing is a failure with a Retry, not a success."""
        scripted('<tool name="run_python">{"code": "print(1)"')
        tid = run_turn(conversation, "compute something")
        assert wait_for_turn(tid) == "failed"

        row = db.connect().execute(
            "SELECT error_code, error_retryable FROM turns WHERE id=?", (tid,)).fetchone()
        assert row["error_code"] == "empty_reply"
        assert row["error_retryable"] == 1

    def test_a_silent_execution_says_so_instead_of_looking_successful(
            self, conversation, scripted, sandbox_ready):
        """REPL-style code prints nothing. Reported as a bland success, the
        model filled the silence with a number it invented — 61,013 for a value
        that is 61,513."""
        scripted(
            '<tool name="run_python">\n{"code": "result = 137 * 449; result"}\n</tool>',
            "I should not guess.",
        )
        tid = run_turn(conversation, "multiply 137 by 449")
        assert wait_for_turn(tid, timeout=180) == "completed"

        session = sandbox.get(sandbox.for_turn(tid)[0]["id"])
        assert session["exit_code"] == 0
        assert not (session["stdout"] or "").strip(), "this test needs a silent run"

        from primnox2.tools import builtins, runtime
        from primnox2.tools.registry import ToolContext

        result = builtins.__dict__["_execute_code"](
            "python", "result = 1 + 1; result", ToolContext(), "safe")
        told = runtime.format_result(result)
        assert "no output" in told.lower()
        assert "print(" in told, "the model is not told how to fix it"
        assert "not seen" in told, "the model is not told to stop guessing"


class TestCodePayloadForms:
    """JSON is a hostile envelope for source code, which is the main thing the
    tool protocol carries. Every string here came off qwen2.5:7b."""

    def test_code_containing_quotes_survives(self, conversation, scripted,
                                              sandbox_ready):
        """The failure that made every document task fail while arithmetic
        passed: `strftime("%Y-%m-%d")` closes the JSON string it sits in."""
        broken = (
            '<tool name="run_python">\n'
            '{"code": "import datetime\nstamp = datetime.date(2026, 8, 14).strftime("%Y-%m-%d")\nprint(stamp)"}\n'
            "</tool>"
        )
        scripted(broken, "Done.")
        tid = run_turn(conversation, "print a date")
        assert wait_for_turn(tid, timeout=180) == "completed"

        sessions = sandbox.for_turn(tid)
        assert sessions, "the call was discarded rather than salvaged"
        assert "2026-08-14" in (sandbox.get(sessions[0]["id"])["stdout"] or "")

    def test_a_fenced_body_needs_no_escaping_at_all(self, conversation, scripted,
                                                    sandbox_ready):
        scripted(
            '<tool name="run_python">\n```python\nprint("quotes \'and\' apostrophes")\n```\n</tool>',
            "Done.",
        )
        tid = run_turn(conversation, "print something quoted")
        assert wait_for_turn(tid, timeout=180) == "completed"

        session = sandbox.get(sandbox.for_turn(tid)[0]["id"])
        assert "quotes 'and' apostrophes" in (session["stdout"] or "")

    def test_the_code_that_ran_is_recorded(self, conversation, scripted,
                                           sandbox_ready):
        """Sessions kept stdout and the file diff but not the source, so a
        surprising result could not be traced to its cause."""
        scripted('<tool name="run_python">\n```\nprint(6 * 7)\n```\n</tool>', "42.")
        tid = run_turn(conversation, "multiply")
        assert wait_for_turn(tid, timeout=180) == "completed"

        session = sandbox.get(sandbox.for_turn(tid)[0]["id"])
        assert "print(6 * 7)" in (session["code"] or ""), "the source was not kept"

    def test_a_multi_argument_tool_still_requires_json(self, conversation):
        """The raw-body form is only safe where the envelope adds nothing. A
        tool with three arguments needs the structure."""
        from primnox2.tools import runtime

        call = runtime.parse_call(
            '<tool name="create_workspace">\nnot json at all\n</tool>')
        assert call is not None
        assert call.get("malformed"), "a multi-argument tool accepted a raw body"


class TestFailedToolGuidance:
    def test_a_failing_tool_says_how_to_retry(self):
        """Measured: after a run failed, the model wrote 'let me correct this'
        and put the fix in a markdown fence in its reply. Prose is deliberately
        never executed, so the turn produced nothing at all."""
        from primnox2.tools import runtime

        told = runtime.format_result({
            "type": "tool_result", "tool": "run_python", "status": "error",
            "summary": "NameError: name 'random' is not defined", "output": "",
        })
        assert "<tool name=\"run_python\">" in told
        assert "not run" in told

    def test_a_successful_tool_is_not_nagged(self):
        from primnox2.tools import runtime

        told = runtime.format_result({
            "type": "tool_result", "tool": "run_python", "status": "success",
            "summary": "ran successfully", "output": "42",
        })
        assert "retry" not in told.lower()


class TestAssetPreviews:
    """The built-in viewers. Every format Primnox can produce should be
    readable without leaving the app, and none of it may be writable."""

    def _asset(self, conversation, data: bytes, name: str) -> dict:
        from primnox2.assets import service as assets_service

        a = assets_service.ingest_bytes(data, name, conversation_id=conversation)
        wait_until(lambda: assets_service.get(a["id"])["status"] == "ready",
                   what=f"{name} to be ingested")
        return assets_service.get(a["id"])

    def test_a_spreadsheet_is_read_despite_its_stored_name(self, conversation):
        """Assets are stored content-addressed, so the file on disk is named
        after its sha256 with no extension — and openpyxl decides the format
        from the filename. Handed the path, it refused a perfectly good
        workbook."""
        import io as _io

        import openpyxl
        from primnox2.assets import preview

        wb = openpyxl.Workbook()
        wb.active.title = "Data"
        wb.active.append(["Region", "Revenue"])
        wb.active.append(["North", 120])
        second = wb.create_sheet("Summary")
        second.append(["Region", "Total"])
        second.append(["North", 120])
        buf = _io.BytesIO()
        wb.save(buf)

        p = preview.describe(self._asset(conversation, buf.getvalue(), "books.xlsx"))
        assert p["kind"] == "sheets"
        assert [s["name"] for s in p["sheets"]] == ["Data", "Summary"]
        assert p["sheets"][0]["header"] == ["Region", "Revenue"]
        assert p["sheets"][0]["rows"] == [["North", "120"]]

    def test_a_csv_sheet_is_named_after_the_file_not_its_hash(self, conversation):
        from primnox2.assets import preview

        p = preview.describe(self._asset(
            conversation, b"a,b\n1,2\n3,4\n", "readings.csv"))
        assert p["kind"] == "sheets"
        assert p["sheets"][0]["name"] == "readings", \
            "the sheet was labelled with the content-address, not the filename"
        assert p["sheets"][0]["header"] == ["a", "b"]
        assert p["sheets"][0]["rows"] == [["1", "2"], ["3", "4"]]

    def test_pdfs_and_images_are_left_to_the_browser(self, conversation):
        from primnox2.assets import preview

        pdf = preview.describe(self._asset(conversation, b"%PDF-1.4\n%%EOF\n", "x.pdf"))
        assert pdf["kind"] == "pdf"
        assert "text" not in pdf, "the bytes were copied through needlessly"

    def test_an_unknown_format_says_so_rather_than_breaking(self, conversation):
        from primnox2.assets import preview

        p = preview.describe(self._asset(conversation, b"\x00\x01binary", "thing.bin"))
        assert p["kind"] in ("unsupported", "text")

    def test_a_corrupt_file_reports_why(self, conversation):
        """A file that cannot be parsed is still a file worth downloading."""
        from primnox2.assets import preview

        p = preview.describe(self._asset(conversation, b"not a workbook", "broken.xlsx"))
        assert p["kind"] == "unreadable"
        assert p["error"]

    def test_the_preview_layer_has_no_write_path(self):
        """Read-only is meant structurally, not by convention."""
        import inspect

        from primnox2.assets import preview

        source = inspect.getsource(preview)
        for forbidden in ("write_bytes(", "write_text(", "wb.save(", ".commit("):
            assert forbidden not in source, f"the preview layer calls {forbidden}"


class TestConversationManagement:
    """Rename, pin, file and remove. The columns were in the schema from the
    start — `folder_id`, `archived_at`, the `folders` table — and nothing had
    ever written to them."""

    def test_renaming_takes_and_survives_a_read(self, conversation):
        turns.rename_conversation(conversation, "  Trip planning  ")
        listed = {c["id"]: c for c in turns.list_conversations()}
        assert listed[conversation]["title"] == "Trip planning", "not trimmed or not saved"

    def test_an_empty_title_is_refused(self, conversation):
        with pytest.raises(ValueError):
            turns.rename_conversation(conversation, "   ")

    def test_pinned_conversations_come_first_and_keep_their_order(self):
        first = turns.create_conversation("pin one")["id"]
        second = turns.create_conversation("pin two")["id"]
        turns.create_conversation("not pinned")

        turns.set_pinned(first, True)
        turns.set_pinned(second, True)
        # Touching the second must not reorder the pins — that is the whole
        # reason pinning is a timestamp rather than a flag.
        turns.rename_conversation(second, "pin two, renamed")

        order = [c["id"] for c in turns.list_conversations()]
        assert order.index(first) < order.index(second), \
            "pins are not in the order they were pinned"
        assert all(order.index(p) < order.index(u)
                   for p in (first, second)
                   for u in [c["id"] for c in turns.list_conversations()
                             if not c.get("pinned_at")]), "a pin sorted below an unpinned chat"

        turns.set_pinned(first, False)
        turns.set_pinned(second, False)

    def test_unpinning_returns_it_to_the_ordinary_list(self):
        """Position is compared against a control rather than against index 0:
        the suite shares one database, so another test's pin can legitimately
        sit above this one.

        `cid` is created FIRST so the control is genuinely newer. With the
        original order the control was older, so recency and pinning both put
        `cid` on top and the post-unpin assertion could only pass when the two
        creations collided inside one millisecond and an untied ORDER BY handed
        them back in insertion order. It did, until the suite got slow enough to
        straddle a millisecond, at which point it failed about one run in five.
        """
        cid = turns.create_conversation("temporary pin")["id"]
        control = turns.create_conversation("unpinned control")["id"]

        turns.set_pinned(cid, True)
        order = [c["id"] for c in turns.list_conversations()]
        assert order.index(cid) < order.index(control), \
            "a pinned conversation sorted below a newer unpinned one"

        turns.set_pinned(cid, False)
        order = [c["id"] for c in turns.list_conversations()]
        assert order.index(cid) > order.index(control), \
            "unpinning did not return it to recency order"

    def test_ordering_is_stable_when_timestamps_collide(self):
        """Conversations created in the same millisecond must not reorder
        between two identical reads.

        `updated_at` is millisecond-resolution, so a burst of creations ties,
        and an untied ORDER BY lets SQLite return tied rows in any order it
        likes — a sidebar that reshuffles on refresh. Pins tie the same way on
        `pinned_at`, and they resolve in the OPPOSITE direction from unpinned
        chats, which is what makes this worth pinning down in a test.
        """
        made = [turns.create_conversation(f"burst {i}")["id"] for i in range(6)]
        for cid in made[:3]:
            turns.set_pinned(cid, True)

        reads = [[c["id"] for c in turns.list_conversations()] for _ in range(5)]
        assert all(r == reads[0] for r in reads), "listing reordered between reads"

        order = reads[0]
        assert [c for c in order if c in made[:3]] == made[:3], \
            "tied pins are not in the order they were pinned"
        assert [c for c in order if c in made[3:]] == list(reversed(made[3:])), \
            "tied unpinned chats are not newest-first"

        for cid in made[:3]:
            turns.set_pinned(cid, False)

    def test_archiving_hides_without_destroying(self):
        cid = turns.create_conversation("to archive")["id"]
        run = turns.create_turn(cid, "something worth keeping")
        turns.archive_conversation(cid)

        assert cid not in [c["id"] for c in turns.list_conversations()]
        assert cid in [c["id"] for c in turns.list_conversations(archived=True)]
        assert turns.get_history(cid), "archiving destroyed the transcript"
        assert turns.status_of(run["turn_id"]) is not None

        turns.archive_conversation(cid, False)
        assert cid in [c["id"] for c in turns.list_conversations()]

    def test_deleting_removes_the_conversation_and_its_turns(self):
        cid = turns.create_conversation("to delete")["id"]
        tid = turns.create_turn(cid, "goodbye")["turn_id"]
        turns.delete_conversation(cid)

        assert db.connect().execute(
            "SELECT 1 FROM conversations WHERE id=?", (cid,)).fetchone() is None
        assert db.connect().execute(
            "SELECT 1 FROM turns WHERE id=?", (tid,)).fetchone() is None

    def test_deleting_a_conversation_keeps_the_documents_it_produced(self):
        """A report should not vanish because the chat that made it was tidied
        away (§2.1 — a conversation owns turns and nothing else)."""
        cid = turns.create_conversation("produced something")["id"]
        tid = turns.create_turn(cid, "make a file")["turn_id"]
        asset = assets.ingest_bytes(b"the report", "report.txt",
                                    conversation_id=cid, turn_id=tid)
        wait_until(lambda: assets.get(asset["id"])["status"] == "ready", what="ingest")

        turns.delete_conversation(cid)
        assert assets.get(asset["id"]) is not None, "the document was destroyed with the chat"

    def test_deleting_an_unknown_conversation_is_an_error_not_a_silence(self):
        with pytest.raises(KeyError):
            turns.delete_conversation("conv_does_not_exist")

    def test_folders_hold_conversations_and_outlive_none_of_them(self):
        folder = turns.create_folder("Research")
        cid = turns.create_conversation("filed")["id"]
        turns.move_conversation(cid, folder["id"])

        listed = {c["id"]: c for c in turns.list_conversations()}
        assert listed[cid]["folder_id"] == folder["id"]
        assert [f for f in turns.list_folders()
                if f["id"] == folder["id"]][0]["conversation_count"] == 1

        # Deleting the folder must not take the conversation with it.
        turns.delete_folder(folder["id"])
        listed = {c["id"]: c for c in turns.list_conversations()}
        assert cid in listed, "deleting a folder deleted its conversations"
        assert listed[cid]["folder_id"] is None

    def test_moving_to_a_folder_that_does_not_exist_is_refused(self):
        cid = turns.create_conversation("stray")["id"]
        with pytest.raises(KeyError):
            turns.move_conversation(cid, "folder_nope")

    def test_an_incognito_conversation_can_be_renamed_but_not_pinned(self):
        from primnox2.chat import ephemeral

        try:
            conv = turns.create_conversation("Private", incognito=True)
            turns.rename_conversation(conv["id"], "Still private")
            assert [c for c in turns.list_conversations()
                    if c["id"] == conv["id"]][0]["title"] == "Still private"
            # Pinning promises permanence that a RAM-only conversation cannot
            # keep past the next restart.
            with pytest.raises(ValueError):
                turns.set_pinned(conv["id"], True)
        finally:
            ephemeral.reset()
